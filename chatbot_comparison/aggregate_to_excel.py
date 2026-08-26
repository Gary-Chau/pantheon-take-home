"""Aggregates every per-(task, model) JSON file in result/raw/ into a single
Excel workbook for manual scoring and analysis.

Usage:
    python aggregate_to_excel.py
"""
from __future__ import annotations

import glob
import json
import os

import hydra
import pandas as pd
from omegaconf import DictConfig


def load_records(raw_result_dir: str) -> pd.DataFrame:
    paths = sorted(glob.glob(os.path.join(raw_result_dir, "*.json")))
    if not paths:
        raise SystemExit(f"No result files found in {raw_result_dir}/. Run run_comparison.py first.")

    rows = []
    for path in paths:
        with open(path, "r", encoding="utf-8") as f:
            record = json.load(f)

        meta = record.get("ollama_metadata") or {}
        eval_count = meta.get("eval_count")
        eval_duration_ns = meta.get("eval_duration_ns")
        tokens_per_second = (
            eval_count / (eval_duration_ns / 1e9)
            if eval_count and eval_duration_ns
            else None
        )

        # Full prompt text, concatenated, as the "complete prompt" record required by the assignment
        prompt_text = "\n\n".join(
            f"[{m['role'].upper()}]\n{m['content']}" for m in record["messages"]
        )

        rows.append({
            "task_id": record["task_id"],
            "category": record["category"],
            "technique_notes": record.get("technique_notes", ""),
            "model_family": record["model_family"],
            "model": record["model_display_name"],
            "model_tag": record["model_tag"],
            "full_prompt": prompt_text,
            "response": record["response"],
            "response_word_count": len(record["response"].split()) if record["response"] else 0,
            "generation_options": json.dumps(record["generation_options"]),
            "thinking_enabled": record.get("think"),
            "total_duration_sec": (meta.get("total_duration_ns") or 0) / 1e9,
            "eval_tokens": eval_count,
            "tokens_per_second": tokens_per_second,
            "wall_clock_seconds": meta.get("wall_clock_seconds"),
            "error": record.get("error"),
            "timestamp_utc": record["timestamp_utc"],
            # Empty columns for manual human scoring, per the assignment's comparison axes
            "score_content_quality_1to5": None,
            "score_contextual_understanding_1to5": None,
            "score_language_fluency_1to5": None,
            "score_ethical_considerations_1to5": None,
            "reviewer_notes": "",
        })

    return pd.DataFrame(rows)


def write_excel(df: pd.DataFrame, path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.sort_values(["task_id", "model"]).to_excel(writer, sheet_name="All Responses", index=False)

        for task_id, group in df.groupby("task_id"):
            sheet_name = task_id[:31]  # Excel sheet name limit
            pivot = group[[
                "model", "full_prompt", "response", "response_word_count",
                "tokens_per_second", "total_duration_sec", "error",
            ]]
            pivot.to_excel(writer, sheet_name=sheet_name, index=False)

        summary = (
            df.groupby("model")
            .agg(
                avg_response_word_count=("response_word_count", "mean"),
                avg_tokens_per_second=("tokens_per_second", "mean"),
                avg_total_duration_sec=("total_duration_sec", "mean"),
                n_errors=("error", lambda s: s.notna().sum()),
            )
            .reset_index()
        )
        summary.to_excel(writer, sheet_name="Summary Stats", index=False)

    _autofit_columns(path)


def _autofit_columns(path: str) -> None:
    """Widens columns and enables text wrap so prompts/responses are readable."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    for sheet in wb.worksheets:
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            header = str(col_cells[0].value or "")
            width = 60 if header in {"full_prompt", "response", "reviewer_notes"} else min(max(len(header) + 2, 12), 30)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
            for cell in col_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


@hydra.main(config_path="configs", config_name="config")
def main(cfg: DictConfig) -> None:
    df = load_records(cfg.raw_result_dir)
    write_excel(df, cfg.excel_output_path)
    print(f"Wrote {len(df)} rows across {df['task_id'].nunique()} tasks and {df['model'].nunique()} models")
    print(f"-> {cfg.excel_output_path}")


if __name__ == "__main__":
    main()
